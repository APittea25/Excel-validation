import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import tempfile

st.title("📊 Cashflow Model Validator")
st.write("Upload your actuarial cashflow Excel file to verify calculations and review formula logic.")

# Upload file
st.markdown("### 📥 Upload Current and Previous Excel Files")
col1, col2 = st.columns(2)

with col1:
    uploaded_file = st.file_uploader("Upload current version", type=[".xlsx"], key="current")

with col2:
    previous_file = st.file_uploader("Upload previous version", type=[".xlsx"], key="previous")


if uploaded_file:
    if previous_file:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_prev:
            tmp_prev.write(previous_file.read())
            prev_path = tmp_prev.name

        df_prev = pd.read_excel(prev_path, sheet_name=0)

        st.subheader("🔍 Input Comparison (Current vs Previous)")
        comparison_inputs = ['Cashflow', 'Death rate', 'Discount rate']

        st.markdown("### 🚨 Anomaly Detection")
        anomalies = []
        for col in comparison_inputs:
            if df[col].isnull().any():
                anomalies.append(f"Missing values detected in {col}.")
            jumps = df[col].diff().abs()
            if (jumps > jumps.mean() + 3 * jumps.std()).any():
                anomalies.append(f"Unusual jump detected in {col} at one or more time steps.")

        if anomalies:
            st.error("Anomalies Detected:")
            for issue in anomalies:
                st.write(f"- {issue}")
        else:
            st.success("No anomalies detected in inputs.")

        input_comparison = pd.DataFrame({'Time': df['Time']})
        for col in comparison_inputs:
            input_comparison[f'{col} (Previous)'] = df_prev[col]
            input_comparison[f'{col} (Current)'] = df[col]
            input_comparison[f'{col} (% Change)'] = 100 * (df[col] - df_prev[col]) / df_prev[col]

        def highlight_large_changes(val):
            try:
                return 'background-color: #ffcccc' if abs(val) > 10 else ''
            except:
                return ''

        styled_comparison = input_comparison.style.applymap(highlight_large_changes, subset=[
            'Cashflow (% Change)', 'Death rate (% Change)', 'Discount rate (% Change)'])

        st.dataframe(styled_comparison)

        st.markdown("### 🧠 AI Summary of Changes")
        import openai
        import os

        openai.api_key = os.getenv("OPENAI_API_KEY")

        summary_prompt = f"""
You are an actuarial analyst reviewing assumption updates in a cashflow model. Summarize the changes in inputs below.

Cashflow (% Change):
{input_comparison['Cashflow (% Change)'].round(2).to_list()}

Death rate (% Change):
{input_comparison['Death rate (% Change)'].round(2).to_list()}

Discount rate (% Change):
{input_comparison['Discount rate (% Change)'].round(2).to_list()}

Write a concise summary highlighting any trends, spikes, or anomalies.
"""

        try:
            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": "You are an actuary skilled at summarizing changes in financial model inputs."},
                    {"role": "user", "content": summary_prompt}
                ]
            )
            ai_comment = response.choices[0].message['content']
            st.success("AI-Generated Summary:")
            st.write(ai_comment)
        except Exception as e:
            st.error(f"OpenAI API error: {e}")

    # Proceed with current file processing
    required_columns = {'Time', 'Cashflow', 'Death rate', 'Discount rate', 'Survival rate', 'Discount factor', 'Expected Cashflow', 'Discounted cashflow', 'PVFP'}
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name

        df = pd.read_excel(tmp_path, sheet_name=0)
    missing_columns = required_columns - set(df.columns)
    if missing_columns:
        st.error(f"Missing required columns in the uploaded file: {', '.join(missing_columns)}")
        st.stop()
    st.subheader("Raw Data Preview")
    st.dataframe(df)

    # Load workbook for formula analysis
    wb = load_workbook(tmp_path, data_only=False)
    ws = wb.active

    # Extract headers
    headers = [cell.value for cell in ws[1]]
    column_formulas = {header: [] for header in headers if header is not None}

    for col_idx, header in enumerate(headers, start=1):
        if header is None:
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            column_formulas[header].append(cell.value)

    column_analysis = {}
    column_descriptions = {
        "Time": "Represents the time period (e.g., year). This is an input.",
        "Cashflow": "Cash amount expected or paid at each time step. This is an input.",
        "Death rate": "Assumed probability of death in the period. This is an input",
        "Discount rate": "Base rate for discounting future values. Often constant or based on a curve. This is an input",
        "Survival rate": "Calculated as the previous survival rate multiplied by (1 - death rate). The survival rate is assumed to be 1 at Time = 1",
        "Discount factor": "Calculated as the previous discount factor divided by (1 + Discount rate). The discount factor is assumed to be 1 at Time = 1",
        "Expected Cashflow": "Calculated as Cashflow × Survival rate.",
        "Discounted cashflow": "Calculated as Expected Cashflow × Discount Factor.",
        "PVFP": "Present Value of Future Profits. =SUM of discounted cashflows.",
    }

    for header, values in column_formulas.items():
        formulas = [v for v in values if isinstance(v, str) and v.startswith('=')]
        description = column_descriptions.get(header, "No specific description available.")
        if not formulas:
            column_analysis[header] = f"✅ Hardcoded values. {description}"
        else:
            unique_formulas = set(formulas)
            if len(unique_formulas) == 1:
                column_analysis[header] = f"❗ Formula-driven: `{formulas[0]}`. {description}"
            else:
                column_analysis[header] = f"❗ Formula-driven (varied formulas). {description}"

    with st.expander("🔍 Column-by-Column Explanation"):
        for col, explanation in column_analysis.items():
            st.write(f"**{col}**: {explanation}")

    with st.expander("Validation Results"):
        errors = []

        # --- Recalculations ---
        df['Discount factor (calc)'] = 1.0
        for i in range(1, len(df)):
            df.loc[i, 'Discount factor (calc)'] = df.loc[i-1, 'Discount factor (calc)'] / (1 + df.loc[i, 'Discount rate'])

        df['Survival rate (calc)'] = 1.0
        for i in range(1, len(df)):
            df.loc[i, 'Survival rate (calc)'] = df.loc[i-1, 'Survival rate (calc)'] * (1 - df.loc[i, 'Death rate'])

        df['Expected Cashflow (calc)'] = df['Cashflow'] * df['Survival rate (calc)']
        df['Discounted Cashflow (calc)'] = df['Expected Cashflow (calc)'] * df['Discount factor']
        df['PVFP (calc)'] = df['Discounted Cashflow (calc)'].sum()

        df['Discount factor diff'] = abs(df['Discount factor'] - df['Discount factor (calc)'])

# --- Check differences ---
        df['Survival rate diff'] = abs(df['Survival rate'] - df['Survival rate (calc)'])
        df['Expected CF diff'] = abs(df['Expected Cashflow'] - df['Expected Cashflow (calc)'])
        df['Discounted CF diff'] = abs(df['Discounted cashflow'] - df['Discounted Cashflow (calc)'])

        tol = 1e-6  # tolerance for float comparison

        if any(df['Survival rate diff'] > tol):
            errors.append("Survival rate calculation mismatch detected.")
        if any(df['Discount factor diff'] > tol):
            errors.append("Discount factor calculation mismatch detected.")

        if any(df['Expected CF diff'] > tol):
            errors.append("Expected Cashflow mismatch detected.")
        if any(df['Discounted CF diff'] > tol):
            errors.append("Discounted Cashflow mismatch detected.")

        pvfp_sheet = df.loc[0, 'PVFP'] if 'PVFP' in df.columns else None
        if pvfp_sheet is not None:
            if abs(df['PVFP (calc)'].iloc[0] - pvfp_sheet) > tol:
                errors.append(f"PVFP mismatch. Calculated: {df['PVFP (calc)'].iloc[0]:.2f}, Sheet: {pvfp_sheet:.2f}")

        # --- Output ---
        if not errors:
            st.success("All calculations are correct. ✅")
        else:
            st.error("Issues found in the following areas:")
            for err in errors:
                st.write(f"- {err}")

        st.subheader("Validation Tables")

        st.markdown("#### Survival Rate")
        st.markdown("""**Calculation Description:** Survival rate is calculated as the previous period's survival rate multiplied by (1 - death rate). The first value is set to 1.0.

```python
df['Survival rate (calc)'] = 1.0
for i in range(1, len(df)):
    df.loc[i, 'Survival rate (calc)'] = df.loc[i-1, 'Survival rate (calc)'] * (1 - df.loc[i, 'Death rate'])
```""")
        st.dataframe(df[['Time', 'Survival rate', 'Survival rate (calc)', 'Survival rate diff']])

        st.markdown("#### Discount Factor")
        st.markdown("""**Calculation Description:** The discount factor is initialized at 1.0. For each subsequent period, it is calculated as 1 / (1 + previous period's discount rate).

```python
df['Discount factor (calc)'] = 1.0
for i in range(1, len(df)):
    df.loc[i, 'Discount factor (calc)'] = df.loc[i-1, 'Discount factor (calc)'] / (1 + df.loc[i, 'Discount rate'])
```""")
        st.dataframe(df[['Time', 'Discount factor', 'Discount factor (calc)', 'Discount factor diff']])

        

        
        st.markdown("#### Expected Cashflow")
        st.markdown("""**Calculation Description:** Expected Cashflow is calculated by multiplying the Cashflow by the Survival rate.

```python
df['Expected Cashflow (calc)'] = df['Cashflow'] * df['Survival rate (calc)']
```""")
        st.dataframe(df[['Time', 'Expected Cashflow', 'Expected Cashflow (calc)', 'Expected CF diff']])

        st.markdown("#### Discounted Cashflow")
        st.markdown("""**Calculation Description:** Discounted Cashflow is calculated by multiplying the Expected Cashflow by the Discount factor.

```python
df['Discounted Cashflow (calc)'] = df['Expected Cashflow (calc)'] * df['Discount factor']
```""")
        st.dataframe(df[['Time', 'Discounted cashflow', 'Discounted Cashflow (calc)', 'Discounted CF diff']])

        if 'PVFP' in df.columns:
            st.markdown("#### PVFP")
            st.markdown("""**Calculation Description:** PVFP (Present Value of Future Profits) is the total of all discounted cashflows.

```python
df['PVFP (calc)'] = df['Discounted Cashflow (calc)'].sum()
```""")
            pvfp_df = pd.DataFrame({
                'PVFP (Excel)': [pvfp_sheet],
                'PVFP (Calculated)': [df['PVFP (calc)'].iloc[0]],
                'Difference': [abs(df['PVFP (calc)'].iloc[0] - pvfp_sheet)]
            })
            st.dataframe(pvfp_df)
